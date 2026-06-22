"""
Extract complex, misaligned, or scanned PDF tables with a multimodal API.

Default target folder:
    C:\Users\Spica\Desktop\W同位素 表生

Install dependencies:
    pip install openai pymupdf openpyxl pillow

Set your API key before running:
    PowerShell:
        $env:OPENAI_API_KEY="sk-..."
    CMD:
        set OPENAI_API_KEY=sk-...

Example:
    python extract_complex_pdf_tables_with_api.py --pdf "paper1.pdf"
    python extract_complex_pdf_tables_with_api.py --limit 3 --model gpt-4o

Important:
    This script uploads rendered PDF page images to the model provider. Do not
    use it for unpublished, confidential, or review-only material unless your
    data policy allows that upload.
"""

from __future__ import annotations

import argparse
import base64
import json
import os
import re
import time
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import fitz  # PyMuPDF
from openai import OpenAI
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


DEFAULT_INPUT_DIR = Path(r"C:\Users\Spica\Desktop\W同位素 表生")
DEFAULT_OUTPUT_DIRNAME = "api_xlsx_outputs"
DEFAULT_MODEL = os.environ.get("OPENAI_TABLE_MODEL", "gpt-4o")

TABLE_ID_RE = re.compile(r"\b(?:Table|TABLE)\s*(?:S\s*)?\d+[A-Za-z]?\b|\b表\s*\d+[A-Za-z]?\b")


@dataclass
class ApiTable:
    table_id: str
    caption: str = ""
    rows: list[list[str]] = field(default_factory=list)
    footnotes: list[str] = field(default_factory=list)
    pages: list[int] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    confidence: str = "unknown"


def normalize_cell(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\x00", "")
    text = text.replace("−", "-") if False else text  # Keep original minus-like symbols by default.
    text = re.sub(r"[ \t]+", " ", text)
    return text.strip()


def normalize_rows(rows: Any) -> list[list[str]]:
    if not isinstance(rows, list):
        return []
    normalized: list[list[str]] = []
    for row in rows:
        if not isinstance(row, list):
            continue
        normalized.append([normalize_cell(cell) for cell in row])
    max_cols = max((len(row) for row in normalized), default=0)
    return [row + [""] * (max_cols - len(row)) for row in normalized]


def safe_sheet_name(name: str, used: set[str]) -> str:
    clean = re.sub(r"[\\/*?:\[\]]", "_", name).strip() or "Table"
    clean = clean[:31]
    if clean not in used:
        used.add(clean)
        return clean
    base = clean[:28]
    index = 2
    while True:
        candidate = f"{base}_{index}"[:31]
        if candidate not in used:
            used.add(candidate)
            return candidate
        index += 1


def data_url_from_png(path: Path) -> str:
    encoded = base64.b64encode(path.read_bytes()).decode("ascii")
    return f"data:image/png;base64,{encoded}"


def render_pdf_pages(pdf_path: Path, image_dir: Path, dpi: int) -> list[Path]:
    image_dir.mkdir(parents=True, exist_ok=True)
    doc = fitz.open(pdf_path)
    images: list[Path] = []
    zoom = dpi / 72
    matrix = fitz.Matrix(zoom, zoom)
    for page_index in range(doc.page_count):
        page = doc.load_page(page_index)
        pix = page.get_pixmap(matrix=matrix, alpha=False)
        out = image_dir / f"page_{page_index + 1:04d}.png"
        pix.save(out)
        images.append(out)
    doc.close()
    return images


def strip_json_fences(text: str) -> str:
    text = text.strip()
    if text.startswith("```"):
        text = re.sub(r"^```(?:json)?\s*", "", text)
        text = re.sub(r"\s*```$", "", text)
    return text.strip()


def parse_json_response(text: str) -> dict[str, Any]:
    cleaned = strip_json_fences(text)
    try:
        return json.loads(cleaned)
    except json.JSONDecodeError:
        start = cleaned.find("{")
        end = cleaned.rfind("}")
        if start >= 0 and end > start:
            return json.loads(cleaned[start : end + 1])
        raise


def build_prompt(pdf_name: str, page_number: int, previous_table_id: str | None) -> str:
    previous_note = previous_table_id or "none"
    return f"""
You are extracting scientific tables from a PDF page image for isotope geochemistry literature.

PDF file: {pdf_name}
Page number: {page_number}
Previous detected table id: {previous_note}

Task:
Detect only real body tables on this page and convert each table to a faithful row/column structure.

Strict rules:
1. Extract only actual table content. Do not extract body paragraphs, figure captions, page headers, page footers, references, or page numbers.
2. Do not translate, rewrite, normalize, or simplify any header, unit, data value, symbol, footnote, n.d., dash, ±, ‰, μ, δ, Δ, isotope notation, or uncertainty.
3. Preserve multi-level headers, unit rows, group header rows, blank cells, footnotes, and visible row order.
4. If this page continues a table from a previous page, set is_continuation=true and continues_table_id to the original table id when inferable.
5. If the table id is visible, use exactly the original id, such as Table 1, Table 2, Table S1. If not visible but this is a continuation, use the continued table id. If no id is inferable, use Unlabeled_page_{page_number}_table_1, etc.
6. If a table is split across the page, still output the visible rows in order.
7. Keep the same number of columns per row when possible. Use empty strings for blank cells.
8. Put table footnotes in footnotes, not inside the main data rows, unless the footnote is visually part of the table body.
9. If you are uncertain about row/column alignment, still provide the best structure and add a warning.

Return JSON only. No Markdown. No explanations outside JSON.
Schema:
{{
  "tables": [
    {{
      "table_id": "Table 1",
      "caption": "original table caption if visible, otherwise empty string",
      "is_continuation": false,
      "continues_table_id": "Table 1 or empty string",
      "rows": [
        ["header cell 1", "header cell 2"],
        ["unit row or data row", "..."]
      ],
      "footnotes": ["original footnote text if visible"],
      "confidence": "high | medium | low",
      "warnings": ["short warning if alignment is uncertain"]
    }}
  ]
}}
""".strip()


def call_model_for_page(
    client: OpenAI,
    model: str,
    image_path: Path,
    pdf_name: str,
    page_number: int,
    previous_table_id: str | None,
    max_retries: int,
) -> dict[str, Any]:
    prompt = build_prompt(pdf_name, page_number, previous_table_id)
    image_url = data_url_from_png(image_path)

    last_error: Exception | None = None
    for attempt in range(1, max_retries + 1):
        try:
            response = client.chat.completions.create(
                model=model,
                temperature=0,
                response_format={"type": "json_object"},
                messages=[
                    {
                        "role": "user",
                        "content": [
                            {"type": "text", "text": prompt},
                            {"type": "image_url", "image_url": {"url": image_url}},
                        ],
                    }
                ],
            )
            content = response.choices[0].message.content or "{}"
            return parse_json_response(content)
        except Exception as exc:  # noqa: BLE001 - keep long batch alive
            last_error = exc
            sleep_seconds = min(30, 2**attempt)
            print(f"    API/page parse failed on attempt {attempt}: {exc}")
            if attempt < max_retries:
                time.sleep(sleep_seconds)
    raise RuntimeError(f"Failed after {max_retries} attempts: {last_error}")


def should_merge_repeated_header(new_rows: list[list[str]], old_rows: list[list[str]]) -> bool:
    if not new_rows or not old_rows:
        return False
    new_header = " | ".join(cell.lower() for cell in new_rows[0] if cell.strip())
    if not new_header:
        return False
    for old_row in old_rows[:5]:
        old_header = " | ".join(cell.lower() for cell in old_row if cell.strip())
        if new_header == old_header:
            return True
    return False


def add_or_merge_table(
    tables: list[ApiTable],
    by_id: dict[str, ApiTable],
    raw: dict[str, Any],
    page_number: int,
    previous_table_id: str | None,
) -> str | None:
    raw_id = normalize_cell(raw.get("table_id", ""))
    continues_id = normalize_cell(raw.get("continues_table_id", ""))
    is_continuation = bool(raw.get("is_continuation", False))

    table_id = raw_id
    if is_continuation and continues_id:
        table_id = continues_id
    elif is_continuation and previous_table_id:
        table_id = previous_table_id
    elif not table_id:
        table_id = f"Unlabeled_page_{page_number}_table_{len(tables) + 1}"

    rows = normalize_rows(raw.get("rows", []))
    if not rows:
        return previous_table_id

    if table_id in by_id:
        table = by_id[table_id]
        if should_merge_repeated_header(rows, table.rows):
            rows = rows[1:]
    else:
        table = ApiTable(table_id=table_id)
        by_id[table_id] = table
        tables.append(table)

    caption = normalize_cell(raw.get("caption", ""))
    if caption and not table.caption:
        table.caption = caption

    table.rows.extend(rows)
    if page_number not in table.pages:
        table.pages.append(page_number)

    footnotes = raw.get("footnotes", [])
    if isinstance(footnotes, list):
        for note in footnotes:
            clean = normalize_cell(note)
            if clean and clean not in table.footnotes:
                table.footnotes.append(clean)

    warnings = raw.get("warnings", [])
    if isinstance(warnings, list):
        for warning in warnings:
            clean = normalize_cell(warning)
            if clean:
                table.warnings.append(f"Page {page_number}: {clean}")

    confidence = normalize_cell(raw.get("confidence", ""))
    if confidence:
        table.confidence = confidence

    return table.table_id


def extract_pdf_with_api(
    client: OpenAI,
    pdf_path: Path,
    output_dir: Path,
    model: str,
    dpi: int,
    include_caption: bool,
    save_page_images: bool,
    max_retries: int,
) -> tuple[list[ApiTable], list[str]]:
    image_root = output_dir / "_page_images" / pdf_path.stem
    page_images = render_pdf_pages(pdf_path, image_root, dpi=dpi)

    tables: list[ApiTable] = []
    by_id: dict[str, ApiTable] = {}
    warnings: list[str] = []
    previous_table_id: str | None = None

    for page_number, image_path in enumerate(page_images, start=1):
        print(f"  Page {page_number}/{len(page_images)}")
        try:
            payload = call_model_for_page(
                client=client,
                model=model,
                image_path=image_path,
                pdf_name=pdf_path.name,
                page_number=page_number,
                previous_table_id=previous_table_id,
                max_retries=max_retries,
            )
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"Page {page_number}: API extraction failed: {exc}")
            continue

        raw_tables = payload.get("tables", [])
        if not isinstance(raw_tables, list):
            warnings.append(f"Page {page_number}: response.tables was not a list")
            continue

        for raw_table in raw_tables:
            if not isinstance(raw_table, dict):
                continue
            previous_table_id = add_or_merge_table(
                tables=tables,
                by_id=by_id,
                raw=raw_table,
                page_number=page_number,
                previous_table_id=previous_table_id,
            )

    if not save_page_images:
        for image in page_images:
            image.unlink(missing_ok=True)
        try:
            image_root.rmdir()
        except OSError:
            pass

    write_workbook(pdf_path, tables, output_dir / f"{pdf_path.stem}.xlsx", include_caption=include_caption)
    return tables, warnings


def guess_header_row_count(rows: list[list[str]]) -> int:
    count = 0
    for row in rows[:6]:
        filled = [cell for cell in row if cell.strip()]
        if not filled:
            continue
        text_like = [cell for cell in filled if re.search(r"[A-Za-z\u4e00-\u9fff]", cell)]
        if len(text_like) / max(1, len(filled)) >= 0.4:
            count += 1
        else:
            break
    return min(max(count, 1), 4) if rows else 0


def autofit_columns(ws, max_width: int = 55) -> None:
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        width = 10
        for cell in ws[letter]:
            if cell.value is None:
                continue
            width = max(width, min(max_width, len(str(cell.value)) + 2))
        ws.column_dimensions[letter].width = width


def write_workbook(pdf_path: Path, tables: list[ApiTable], output_path: Path, include_caption: bool) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    used_names: set[str] = set()
    thin = Side(style="thin", color="D9D9D9")
    header_fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    note_fill = PatternFill(fill_type="solid", fgColor="FFF2CC")

    if not tables:
        ws = wb.create_sheet("No_tables_found")
        ws["A1"] = "No tables were returned by the API. Check page images or rerun with a higher DPI/model."
        wb.save(output_path)
        return

    for table in tables:
        ws = wb.create_sheet(safe_sheet_name(table.table_id, used_names))
        max_cols = max((len(row) for row in table.rows), default=1)
        current_row = 1

        if include_caption and table.caption:
            ws.cell(current_row, 1, table.caption)
            ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
            ws.cell(current_row, 1).font = Font(bold=True)
            ws.cell(current_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
            current_row += 2

        header_rows = guess_header_row_count(table.rows)
        for row_offset, row in enumerate(table.rows):
            excel_row = current_row + row_offset
            for col_idx, value in enumerate(row, start=1):
                cell = ws.cell(excel_row, col_idx, value)
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
                if row_offset < header_rows:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill

        current_row += len(table.rows) + 1

        if table.footnotes:
            ws.cell(current_row, 1, "Footnotes")
            ws.cell(current_row, 1).font = Font(bold=True)
            ws.cell(current_row, 1).fill = note_fill
            current_row += 1
            for note in table.footnotes:
                ws.cell(current_row, 1, note)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
                ws.cell(current_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
                current_row += 1

        if table.warnings:
            current_row += 1
            ws.cell(current_row, 1, "Extraction warnings")
            ws.cell(current_row, 1).font = Font(bold=True)
            ws.cell(current_row, 1).fill = note_fill
            current_row += 1
            for warning in table.warnings:
                ws.cell(current_row, 1, warning)
                ws.merge_cells(start_row=current_row, start_column=1, end_row=current_row, end_column=max_cols)
                ws.cell(current_row, 1).alignment = Alignment(wrap_text=True, vertical="top")
                current_row += 1

        ws.freeze_panes = f"A{1 + header_rows}"
        autofit_columns(ws)

    wb.save(output_path)


def write_summary(summary_rows: list[dict[str, str]], output_dir: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "summary"
    headers = ["PDF", "XLSX", "Tables", "Pages", "Confidence", "Warnings"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill(fill_type="solid", fgColor="EAF2F8")
    for row in summary_rows:
        ws.append([row.get(key, "") for key in headers])
    autofit_columns(ws, max_width=90)
    wb.save(output_dir / "api_extraction_summary.xlsx")


def select_pdfs(input_dir: Path, pdf_names: list[str], limit: int | None) -> list[Path]:
    if pdf_names:
        pdfs = []
        for name in pdf_names:
            path = Path(name)
            if not path.is_absolute():
                path = input_dir / name
            if path.suffix.lower() != ".pdf":
                path = path.with_suffix(".pdf")
            if not path.exists():
                raise FileNotFoundError(path)
            pdfs.append(path)
    else:
        pdfs = sorted(input_dir.glob("*.pdf"))
    return pdfs[:limit] if limit else pdfs


def main() -> None:
    parser = argparse.ArgumentParser(description="Use a multimodal API to extract complex PDF tables to XLSX.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT_DIR, help="Folder containing PDF files.")
    parser.add_argument("--output", type=Path, default=None, help="Output folder for XLSX files.")
    parser.add_argument("--pdf", action="append", default=[], help="Specific PDF filename/path to process. Can be repeated.")
    parser.add_argument("--limit", type=int, default=None, help="Only process the first N PDFs. Useful for testing cost.")
    parser.add_argument("--model", default=DEFAULT_MODEL, help="Vision-capable model name. Default: gpt-4o or OPENAI_TABLE_MODEL.")
    parser.add_argument("--dpi", type=int, default=200, help="Page render DPI. Try 250-300 for tiny tables, but cost increases.")
    parser.add_argument("--no-caption", action="store_true", help="Do not write table captions into the first row.")
    parser.add_argument("--save-page-images", action="store_true", help="Keep rendered page images for manual checking.")
    parser.add_argument("--max-retries", type=int, default=3, help="API retry count per page.")
    args = parser.parse_args()

    if not os.environ.get("OPENAI_API_KEY"):
        raise EnvironmentError("OPENAI_API_KEY is not set. Set it locally before running this script.")

    input_dir = args.input.expanduser().resolve()
    output_dir = args.output.expanduser().resolve() if args.output else input_dir / DEFAULT_OUTPUT_DIRNAME
    output_dir.mkdir(parents=True, exist_ok=True)

    pdfs = select_pdfs(input_dir, args.pdf, args.limit)
    if not pdfs:
        raise FileNotFoundError(f"No PDF files found in: {input_dir}")

    client = OpenAI()
    summary_rows: list[dict[str, str]] = []

    print(f"Model: {args.model}")
    print(f"PDF count: {len(pdfs)}")
    print(f"Output: {output_dir}")

    for index, pdf_path in enumerate(pdfs, start=1):
        print(f"[{index}/{len(pdfs)}] Processing {pdf_path.name}")
        tables, warnings = extract_pdf_with_api(
            client=client,
            pdf_path=pdf_path,
            output_dir=output_dir,
            model=args.model,
            dpi=args.dpi,
            include_caption=not args.no_caption,
            save_page_images=args.save_page_images,
            max_retries=args.max_retries,
        )
        summary_rows.append(
            {
                "PDF": pdf_path.name,
                "XLSX": f"{pdf_path.stem}.xlsx",
                "Tables": "; ".join(table.table_id for table in tables) or "No tables returned",
                "Pages": "; ".join(f"{table.table_id}: p{','.join(map(str, table.pages))}" for table in tables),
                "Confidence": "; ".join(f"{table.table_id}: {table.confidence}" for table in tables),
                "Warnings": " | ".join(warnings + [w for table in tables for w in table.warnings]),
            }
        )

    write_summary(summary_rows, output_dir)
    print(f"Done. Summary: {output_dir / 'api_extraction_summary.xlsx'}")


if __name__ == "__main__":
    main()
